import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


def open_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        print(f"Файл {file_path} успешно открыт")
        return workbook
    except FileNotFoundError:
        print(f"Ошибка: Файл {file_path} не найден")
        raise
    except InvalidFileException:
        print(f"Ошибка: {file_path} не является валидным Excel файлом")
        raise


def get_data_from_table(file_path, search_value):
    try:
        from openpyxl import load_workbook

        normalized_search = search_value.lower().replace(" ", "")

        wb = load_workbook(file_path)

        if "Лист1" not in wb.sheetnames:
            print("Ошибка: лист 'Лист1' не найден в файле")
            return None

        sheet = wb["Лист1"]

        for row in sheet.iter_rows():
            cell_value = str(row[15].value).lower().replace(" ", "") if row[15].value else ""

            if cell_value == normalized_search:
                fact = row[16].value  # Q
                plan = row[17].value  # R
                percent = row[18].value  # S

                if fact is None or plan is None or percent is None:
                    print(f"Предупреждение: не все данные найдены для значения '{search_value}'")
                    return None

                formatted_fact = f"{fact:.2%}".replace(".", ",") if isinstance(fact, (int, float)) else str(fact)
                formatted_plan = f"{plan:.2%}".replace(".", ",") if isinstance(plan, (int, float)) else str(plan)
                formatted_percent = f"{percent:.1%}".replace(".", ",").replace("%", "") if isinstance(percent, (
                int, float)) else str(percent)

                print(f"Найдено: {search_value}")
                print(f"Факт: {formatted_fact}, План: {formatted_plan}, % выполнения: {formatted_percent}")
                return fact, plan, percent

        print(f"Значение '{search_value}' не найдено в столбце P")
        return None

    except Exception as e:
        print(f"Ошибка при обработке файла Excel: {str(e)}")
        return None